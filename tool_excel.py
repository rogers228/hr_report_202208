# excel 相關工具 獨立出來
import os
import openpyxl
from openpyxl.comments import Comment #註解
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D #插入圖片用
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker #插入圖片用
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU #插入圖片用
from openpyxl.drawing.image import Image #插入圖片用
from openpyxl.utils import get_column_letter #轉換

#https://openpyxl.readthedocs.io/en/stable/api/openpyxl.utils.cell.html
# get_column_letter: (3 > C)

from math import ceil

from sys_config import *
from tool_style import *
from tool_math import *


class tool_excel(): #讀取excel 單一零件
    def __init__(self, file, workbook, sh):
        self.file = file
        self.workbook = workbook
        self.sh = sh # excel sheet

    def c_write(self, row, column, value = '', font = font_9, alignment = ah_left, border = no_border, fillcolor = cf_none):
        #寫入儲存格 並設定格式
        cell = self.sh.cell(row, column)
        cell.value = value
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
        if fillcolor:
            cell.fill = fillcolor

    def c_merge(self, start_row, start_column, end_row, end_column):
        self.sh.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column) #合併儲存格

    def c_unmerge(self, start_row, start_column, end_row, end_column):
        self.sh.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column) #取消合併儲存格

    def c_fill(self, row, column, border = no_border, fillcolor = cf_none):
        cell = self.sh.cell(row, column)
        if border:
            cell.border = border
        if fillcolor:
            cell.fill = fillcolor

    def c_comm(self, row, column, comment):
        self.sh.cell(row, column).comment = comment

    def c_line_bottom(self, row, start_column, columns): #畫線 下格線
        for i in range(start_column, start_column + columns):
            self.sh.cell(row, i).border = bottom_border #style

    def c_sh_moto_first(self):
        # 移動到首個位置
        sheetname = self.sh.title
        if not sheetname in self.workbook.sheetnames:
            return

        sheet_index = self.workbook.sheetnames.index(sheetname)
        if sheet_index == 0:
            return

        self.workbook.move_sheet(self.sh, sheet_index * -1) # 移動到首個位置

    def cell_h_to_EMU(h):
        return cm_to_EMU((h * 49.77)/99)

    def cell_w_to_EMU(w):
        return cm_to_EMU((w * (18.65-1.71))/10)

    def c_image(self, row, column, imgPath, width, height, rowoffset=0, coloffset=0): #插入圖片
        # imgPath 圖片路徑  請在程序外先檢查是否存在
        img = Image(imgPath)
        img.width = width
        img.height = height
        cell_h_to_EMU = lambda h: cm_to_EMU((h * 49.77)/99)         # cell height EMU單位
        cell_w_to_EMU = lambda w: cm_to_EMU((w * (18.65-1.71))/10)  # cell width  EMU單位
        coloffset = cell_w_to_EMU(coloffset) #偏移
        rowoffset = cell_h_to_EMU(rowoffset) #偏移
        marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset) #建立標記位置
        size = XDRPositiveSize2D(pixels_to_EMU(width), pixels_to_EMU(height))
        img.anchor = OneCellAnchor(_from=marker, ext=size) #img 定位
        self.sh.add_image(img)

    def c_column_width(self, width_list): # 設定欄寬
        for i in range(len(width_list)):
            self.sh.column_dimensions[get_column_letter(i+1)].width = width_list[i]
            # get_column_letter: (3 > C)

    def set_page(self): # 頁面設定layout
        self.sh.page_margins = openpyxl.worksheet.page.PageMargins(
            left=cm2in(0.7),
            right=cm2in(0.7),
            top=cm2in(1.0),
            bottom=cm2in(0.5),
            header=cm2in(1.0),
            footer=cm2in(0.5))

    def save_xls(self): # 儲存
        try:
            self.workbook.save(self.file) #save
        except:
            print('儲存時發生錯誤，無法處理該檔案，有可能檔案已被開啟尚未關閉!')

    # def get_height_for_row(self, row, column, column_width, md003, font_size=12):
    #     # print('md003', md003)
    #     # print('column_width', column_width)
    #     cell = self.sh.cell(row, column)
    #     font_params = factor_of_font_size_to_width[font_size]
    #     height = font_params["height"]
    #     # words_count_at_one_row = self.sh.column_dimensions[cell.column_letter].width / font_params["factor"]
    #     words_count_at_one_row = column_width / font_params["factor"]
    #     # print('words_count_at_one_row', words_count_at_one_row)
    #     # print('len(str(md003))', len(str(md003)))
    #     lines = ceil(len(str(md003)) / words_count_at_one_row) +2 # 品號 *1 + 品名*1 + 規格自適
    #     # print('lines', lines)
    #     height = max(height, (lines * font_params["height"]) )
    #     return height

    def ger_height_pdinfo(self, s1, s2, s3, column_width_pixel, font_size =9): #品號品名規格所需高度
        font_params = factor_of_font_size_to_width[font_size]
        lines_s1 = self.ger_wrln(s1, column_width_pixel, font_size)
        lines_s2 = self.ger_wrln(s2, column_width_pixel, font_size)
        lines_s3 = self.ger_wrln(s3, column_width_pixel, font_size)
        height = (lines_s1 + lines_s2 + lines_s3) * font_params["height"]
        return height

    def ger_height_onecell(self, mystr, column_width_pixel, font_size =9): #一個儲存格所需需高度
        font_params = factor_of_font_size_to_width[font_size]
        lines = self.ger_wrln(mystr, column_width_pixel, font_size)
        height = lines * font_params["height"]
        return height

    def ger_wrln(self, words, column_width_pixel, font_size =9): #依照文字及寬度計算出所需行數
        font_params = factor_of_font_size_to_width[font_size]
        len_ch = len(list(filter((lambda x:'\u4e00' <= x <= '\u9fa5'), list(words)))) #中文字數
        len_en = len(words)- len_ch                                                   # 其他英數符號字數
        words_length_pixel = len_ch*font_params['one_cher_width_ch_pixel'] + len_en*font_params['one_cher_width_en_pixel'] #文字長度像素
        return max(ceil(words_length_pixel / column_width_pixel),1) # 所需行數 最少一行

    def c_mark_logo(self): # 添加浮水印
        imgPath = "{0}\\{1}".format(s_ImagePath, 'yeoshe_mark_grey.png')
        # print('imgPath',imgPath)
        # 原像素450, 300
        self.c_image(4+8, 2, imgPath, 400, 267, 0, 1.0) #插入圖片

def test1():
    print('test1')

if __name__ == '__main__':
    test1()
    print('ok')