if True:
    import sys, custom_path
    config_path = custom_path.custom_path['hr_report_202208'] # 取得專案引用路徑
    sys.path.append(config_path) # 載入專案路徑

import os
import openpyxl
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D #插入圖片用
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker #插入圖片用
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU #插入圖片用
from openpyxl.drawing.image import Image #插入圖片用
from openpyxl.utils import get_column_letter #轉換
from tool_style import *

class tool_excel(): #讀取excel 單一零件
    def __init__(self, file, workbook, sh):
        self.file = file
        self.workbook = workbook
        self.sh = sh # excel sheet

    def c_write(self, row, column, value = '', font = font_9, alignment = ah_left, border = no_border, fillcolor = cf_none):
        #寫入儲存格 並設定格式
        cell = self.sh.cell(row, column)
        cell.value = value
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
        if fillcolor:
            cell.fill = fillcolor

    def c_merge(self, start_row, start_column, end_row, end_column):
        self.sh.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column) #合併儲存格

    def set_page_layout(self): # 頁面設定layout
        cm2in = lambda x: x/2.54
        self.sh.page_margins = openpyxl.worksheet.page.PageMargins(
            left=cm2in(1.2),
            right=cm2in(1.2),
            top=cm2in(1.0),
            bottom=cm2in(0.5),
            header=cm2in(1.0),
            footer=cm2in(0.5))

    def set_page_layout_horizontal(self): # 頁面設定layout  橫式
        cm2in = lambda x: x/2.54
        self.sh.page_margins = openpyxl.worksheet.page.PageMargins(
            left=cm2in(1.0),
            right=cm2in(1.0),
            top=cm2in(1.0),
            bottom=cm2in(0.5),
            header=cm2in(1.0),
            footer=cm2in(0.5))

        self.sh.page_setup.paperSize = self.sh.PAPERSIZE_A4
        # Paper size 紙張大小
            # PAPERSIZE_LETTER = '1'
            # PAPERSIZE_LETTER_SMALL = '2'
            # PAPERSIZE_TABLOID = '3'
            # PAPERSIZE_LEDGER = '4'
            # PAPERSIZE_LEGAL = '5'
            # PAPERSIZE_STATEMENT = '6'
            # PAPERSIZE_EXECUTIVE = '7'
            # PAPERSIZE_A3 = '8'
            # PAPERSIZE_A4 = '9'
            # PAPERSIZE_A4_SMALL = '10'
            # PAPERSIZE_A5 = '11'
        self.sh.page_setup.orientation = self.sh.ORIENTATION_LANDSCAPE
        # Page orientation 紙張方向
            # ORIENTATION_PORTRAIT = 'portrait' #縱向
            # ORIENTATION_LANDSCAPE = 'landscape' #横向
        self.sh.sheet_view.zoomScale = 100 # 檢視縮放
        self.sh.page_setup.scale = 70      # 列印縮放比例
        self.sh.print_options.horizontalCentered=True # 水平居中

    def c_image(self, row, column, imgPath, width, height, rowoffset=0, coloffset=0): #插入圖片
        # imgPath 圖片路徑  請在程序外先檢查是否存在
        img = Image(imgPath)
        img.width = width
        img.height = height
        cell_h_to_EMU = lambda h: cm_to_EMU((h * 49.77)/99)         # cell height EMU單位
        cell_w_to_EMU = lambda w: cm_to_EMU((w * (18.65-1.71))/10)  # cell width  EMU單位
        coloffset = cell_w_to_EMU(coloffset) #偏移
        rowoffset = cell_h_to_EMU(rowoffset) #偏移
        marker = AnchorMarker(col=column-1, colOff=coloffset, row=row-1, rowOff=rowoffset) #建立標記位置  由1始
        size = XDRPositiveSize2D(pixels_to_EMU(width), pixels_to_EMU(height))
        img.anchor = OneCellAnchor(_from=marker, ext=size) #img 定位
        self.sh.add_image(img)

    def c_column_width(self, width_list): # 設定欄寬
        for i in range(len(width_list)):
            self.sh.column_dimensions[get_column_letter(i+1)].width = width_list[i]

    def c_row_height(self, row_index, row_height): # 設定列高
        self.sh.row_dimensions[row_index].height = row_height 

    def save_xls(self): # 儲存
        try:
            self.workbook.save(self.file) #save
        except:
            print('儲存時發生錯誤，無法處理該檔案，有可能檔案已被開啟尚未關閉!')

    def open_xls(self):
        if os.path.exists(self.file): #檔案存在
            os.startfile(self.file)

def test1():
    print('test1')

if __name__ == '__main__':
    test1()
    print('ok')